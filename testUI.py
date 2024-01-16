
from threading import Thread
import requests,openpyxl,csv,json,sqlite3
from bs4 import BeautifulSoup
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import PatternFill
import tkinter as tk
import tkinter.font as tkFont
from tkinter.filedialog import askopenfilename
from tkinter import ttk,messagebox,Frame, Menu, Toplevel


def search_company_name(afm):
    url = "https://publicity.businessportal.gr/api/search"
    headers={
        "Host":"publicity.businessportal.gr",
        "User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:109.0) Gecko/20100101 Firefox/118.0",
        "Accept":"application/json, text/plain, */*",
        "Accept-Language":"en-US,en;q=0.5",
        "Accept-Encoding":"gzip, deflate, br",
        "Content-Type":"application/json",
        "Content-Length":"496",
        "Origin":"https://publicity.businessportal.gr",
        "Connection":"keep-alive",
        "Referer":"https://publicity.businessportal.gr/"
        }
    data={"dataToBeSent":{"inputField":afm,"city":None,"postcode":None,"legalType":[],"status":[],"suspension":[],"category":[],"specialCharacteristics":[],"employeeNumber":[],"armodiaGEMI":[],"kad":[],"recommendationDateFrom":None,"recommendationDateTo":None,"closingDateFrom":None,"closingDateTo":None,"alterationDateFrom":None,"alterationDateTo":None,"person":[],"personrecommendationDateFrom":None,"personrecommendationDateTo":None,"radioValue":"all","places":[]},"token":None,"language":"el"}
    response = requests.request('post',url, headers=headers,json=data)

    j = json.loads(response.text)
    lst = j['company']['hits']
    no_gemi = ""
    for rec in lst :
        suffix_gemi = rec["id"][-3:]
        if suffix_gemi in "000":
            return rec["name"]

def get_company_name(afm):
    base_url = "https://publicity.businessportal.gr/api/autocomplete/"
    url = base_url+afm

    headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36",
    "Origin": "https://publicity.businessportal.gr",
    "Referer": "https://publicity.businessportal.gr/",
    "Cookie": "next-i18next=el"
    }
    data = {"token":"null","language":"el"}
    response = requests.post(url, headers=headers)
    j = json.loads(response.text)
    lst = j['payload']['autocomplete']
    no_gemi = ""
    for rec in lst :
        if not rec.get("branchType"):
            no_gemi = str(rec['arGemi'])

    url = "https://publicity.businessportal.gr/api/company/details"
    payload = {"query":{"arGEMI":no_gemi},"language":"el"}
    response = requests.post(url, headers=headers,json=payload)
    j2 = json.loads(response.text)
    if j2['message'] in 'Company not found':
        return None
    company_name = j2['companyInfo']['payload']["company"]["name"]
    return company_name

def associate (Receipts,People,Max_Cost):
    N_Receipts = Receipts.copy()

    for p in People:

        for r in N_Receipts:
            if (r.Total > Max_Cost) or (r.Total< 0) or (p.Total + r.Total > Max_Cost) :
                continue
            p.Total += r.Total
            p.Array_Of_receipts.append(r)
            N_Receipts.remove(r)

class Reciept:
    def __init__(self,hmerominia,katharo_sunolo,sunolo_fpa,eidos_parastatikou,sunolo,arithmos_parastatikou,eidos_proiontos,afm,eponumia,url):
        self.Date_time = hmerominia
        self.Sum_of_net_price = katharo_sunolo
        self.Sum_of_fpa = sunolo_fpa
        self.R_type = eidos_parastatikou
        self.Total = float(sunolo.replace("€",""))
        self.Number = arithmos_parastatikou
        self.Type_of_goods = eidos_proiontos
        self.Company_afm = afm
        self.Company_Name = eponumia
        self.AADE_Url = url
    def set_id(self,_id):
        self.ID = _id
    
    def __repr__(self):
        return f"{self.ID}\t{self.Total}"

    def To_tuple(self):
        return (self.ID,self.Total)

    def fix_num(self):
        if type(self.Total) is type("1.1"):
            self.Total = float(self.Total.replace("€",""))

    def convert_to_record(self):
        return (2,self.Date_time,self.Sum_of_net_price,self.Sum_of_fpa," "," ",1," ",self.R_type,self.Number,self.Type_of_goods,"998727941","ΤΑΝΤΕΜ ΑΣΤΙΚΗ ΜΗ ΚΕΡΔΟΣΚΟΠΙΚΗ ΕΤΑΙΡΕΙΑ",1,self.Company_afm,self.Company_Name,"{:.2f}".format(self.Total),"=HYPERLINK(\""+self.AADE_Url+"\")")

class Person:
    def __init__(self,n,aor,tl):
        self.Name = n
        self.Array_Of_receipts = aor
        self.Total = tl
    
    def __repr__(self):
        return f"{self.Name} {self.Array_Of_receipts} {self.Total}"

def OpenFile(this):
    filepath = askopenfilename(filetypes=[("csv Files", "*.csv"), ("All Files", "*.*")])
    if not filepath:
        return None
    with open(filepath,'r',encoding='utf-8') as fl:
        doc = csv.reader(fl)
        for row in doc:
            Urls.append(row[4])

    this.L_File_Path.config(text=f"{filepath}")
    return Urls

def Execute(U,this):
    db = sqlite3.connect("./Database.db")
    cursor = db.cursor()
    Urls.pop(0)
    DATA_ARRAY = []
    errors = []
    Rs = []
    MAX_COST = float(this.Input_Max_amount_of_Money.get()) if this.Input_Max_amount_of_Money.get() != '' else 0
    No_of_candidates = int(this.Input_Number_of_Volunteers.get()) if this.Input_Number_of_Volunteers.get() != '' else 0

    for url in Urls:
        this.progressbar['value'] += 100/len(Urls)
        this.Window.update_idletasks()
        try:
            r = requests.get(url)
        except:
            errors.append(url)
            continue
        soup = BeautifulSoup(r.text,'lxml')
        table = soup.find('table',"info")
        if table is None:
            errors.append(url)
            continue
        tr = table.findAll('tr')
        DictArray = {}
        for row in tr:
            x = row.findAll('td')
            x = x[0].text , x[1].text
            DictArray[x[0]]= x[1]
        try:
            date_time = DictArray["Ημερομηνία, ώρα"].split(" ")[0]
            eidos_parastatikou = DictArray["Είδος παραστατικού"]
            receipt_number = DictArray["Αριθμός παραστατικού"]
            sum_net = str(round(float(DictArray['Καθαρή αξία Α'].replace("€ ","")) + float(DictArray['Καθαρή αξία Β'].replace("€ ","")) + float(DictArray['Καθαρή αξία Γ'].replace("€ ","")) + float(DictArray['Καθαρή αξία Δ'].replace("€ ","")) + float(DictArray['Καθαρή αξία Ε'].replace("€ ","")),2))+"€"
            sum_fpa = str(round(float(DictArray['ΦΠΑ Α'].replace("€ ","")) +float(DictArray['ΦΠΑ Β'].replace("€ ","")) + float(DictArray['ΦΠΑ Γ'].replace("€ ","")) + float(DictArray['ΦΠΑ Δ'].replace("€ ","")),2))+"€"
            sun_axia = str(round(float(DictArray[' Συνολική αξία '].rstrip(" ").replace("€ ","")),2))+"€"
            _afm = DictArray["ΑΦΜ εκδότη"]
        except:
            errors.append(url)
            continue
        cursor.execute(f"select Company_Name from Companies where AFM ='{_afm}';")
        company_name = cursor.fetchone()
        if company_name is not None:
            company_name = company_name[0]
            
        if company_name is None:
            company_name = search_company_name(_afm)
            if company_name is not None:
                cursor.execute(f"insert Into Companies(AFM,Company_name) values(?,?);",(_afm,company_name))

        if company_name is None:
            company_name = get_company_name(_afm)
            if company_name is not None:
                cursor.execute(f"insert Into Companies(AFM,Company_name) values(?,?);",(_afm,company_name))

        cursor.execute(f"select Product_Type from Companies where AFM ='{_afm}';")
        eidos_proiontos = cursor.fetchone()
        if eidos_proiontos is not None:
            eidos_proiontos = eidos_proiontos[0]
        url = url.rstrip('\x00')
        Reciept_Object = Reciept(date_time,sum_net,sum_fpa,eidos_parastatikou,sun_axia,receipt_number,eidos_proiontos,_afm,company_name,url)
        Rs.append(Reciept_Object)
        DATA_ARRAY.append(
            (2,date_time,sum_net,sum_fpa," "," ",1," ",eidos_parastatikou,receipt_number,eidos_proiontos,"998727941","ΤΑΝΤΕΜ ΑΣΤΙΚΗ ΜΗ ΚΕΡΔΟΣΚΟΠΙΚΗ ΕΤΑΙΡΕΙΑ",1,_afm,company_name,sun_axia,"=HYPERLINK(\""+url+"\")"))
    db.commit()


    Hashed_Receipts = {}

    for r in Rs:
        r.set_id(hash(r))
        r.fix_num()
        Hashed_Receipts[r.ID] = r

    items = [(r.ID,r.Total) for r in Rs]

    Rs.sort(key=lambda r:r.Total,reverse=True)

    Person_array = [Person(chr(i+65),[],0) for i in range(0,No_of_candidates)]
    Food_Receipts = [obj for obj in Rs if obj.Type_of_goods=="Τρόφιμα"]
    associate(Food_Receipts,Person_array,MAX_COST)
    Used_Reciepts = []
    for p in Person_array:
        for rec in p.Array_Of_receipts:
            Used_Reciepts.append(rec)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    for row in DATA_ARRAY:
        worksheet.append(row)

    worksheet.append(("","",""))
    for rows in worksheet.iter_rows(min_row=worksheet.max_row, max_row=worksheet.max_row, min_col=None):
        for cell in rows:
            cell.fill = PatternFill(start_color="9aec9f", end_color="9aec9f", fill_type = "solid")
    worksheet.append(("","","Grouped Reciepts","Food"))

    for p in Person_array:
        worksheet.append(("","Volunteer :"+p.Name,"Total","{:.2f}".format(p.Total)))
        for rec in p.Array_Of_receipts:
            worksheet.append(Hashed_Receipts[rec.ID].convert_to_record())
        worksheet.append(("","",""))

    for rows in worksheet.iter_rows(min_row=worksheet.max_row, max_row=worksheet.max_row, min_col=None):
        for cell in rows:
            cell.fill = PatternFill(start_color="619df9", end_color="619df9", fill_type = "solid")
    worksheet.append(("","","Grouped Reciepts","Non-Food"))
    del Person_array

    Person_array = [Person(chr(i+65),[],0) for i in range(0,No_of_candidates)]
    Remaining_Reciepts = [obj for obj in Rs if obj not in Used_Reciepts]
    associate(Remaining_Reciepts,Person_array,MAX_COST)

    for p in Person_array:
        worksheet.append(("","Volunteer :"+p.Name,"Total","{:.2f}".format(p.Total)))
        for rec in p.Array_Of_receipts:
            worksheet.append(Hashed_Receipts[rec.ID].convert_to_record())
        worksheet.append(("","",""))

    for rows in worksheet.iter_rows(min_row=worksheet.max_row, max_row=worksheet.max_row, min_col=None):
        for cell in rows:
            cell.fill = PatternFill(start_color="FF0000", end_color="ff0000", fill_type = "solid")

    worksheet.append((" "," "," "," "," "," "," "," "," "," "," "," "," "," "," "," "," ","ERRORS"))

    for e in errors:
        e = ILLEGAL_CHARACTERS_RE.sub(r'',e)
        worksheet.append((" "," "," "," "," "," "," "," "," "," "," "," "," "," "," "," "," ","=HYPERLINK(\""+e+"\")"))
    workbook.save("OUTPUT.xlsx")
    messagebox.showinfo("showinfo", "Task Completed") 

def Start(this):
    if Urls == []:
        messagebox.showinfo("showinfo", "Please Choose File")
        return None

    t = Thread(target=Execute,args=(Urls,this))
    t.start()
    if not t.is_alive() :
        messagebox.showinfo("showinfo", "Completed click ok to close")
        this.destroy()
    #Execute(Urls)


Urls = []
# region UI inits
class MainWindow:
    def __init__(self,root):
        self.Window = root
        self.Window.title("Recipt to XLS maker")
        width=600
        height=500
        screenwidth = self.Window.winfo_screenwidth()
        screenheight = self.Window.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height, (screenwidth - width) / 2, (screenheight - height) / 2)
        self.Window.geometry(alignstr)
        self.Window.resizable(width=False, height=False)

        self.Window.iconbitmap("./logo.ico")

        MenuBar = Menu(self.Window)
        self.Window.config(menu=MenuBar)

        Menu_File = Menu(MenuBar,tearoff=0)
        Menu_File.add_command(label="Update Database", command=Window_Update_Database)
        Menu_File.add_separator()
        Menu_File.add_command(label="Exit", command=self.Window.destroy)

        MenuBar.add_cascade(label="File", menu=Menu_File)
        MenuBar.add_cascade(label="Help", command=Help_Window)

        self.L_Select_file=tk.Label(self.Window)
        ft = tkFont.Font(family='Times',size=10)
        self.L_Select_file["font"] = ft
        self.L_Select_file["fg"] = "#333333"
        self.L_Select_file["justify"] = "center"
        self.L_Select_file["text"] = "Selected File"
        self.L_Select_file.place(x=60,y=40,width=132,height=32)

        self.L_File_Path=tk.Label(self.Window)
        ft = tkFont.Font(family='Times',size=10)
        self.L_File_Path["font"] = ft
        self.L_File_Path["fg"] = "#333333"
        self.L_File_Path["justify"] = "center"
        self.L_File_Path["text"] = None
        self.L_File_Path.place(x=180,y=40,width=376,height=30)

        L_Number_of_Volunteers=tk.Label(self.Window)
        ft = tkFont.Font(family='Times',size=10)
        L_Number_of_Volunteers["font"] = ft
        L_Number_of_Volunteers["fg"] = "#333333"
        L_Number_of_Volunteers["justify"] = "center"
        L_Number_of_Volunteers["text"] = "Number of Volunteers"
        L_Number_of_Volunteers.place(x=70,y=140,width=171,height=39)

        L_Max_amount_of_Money=tk.Label(self.Window)
        ft = tkFont.Font(family='Times',size=10)
        L_Max_amount_of_Money["font"] = ft
        L_Max_amount_of_Money["fg"] = "#333333"
        L_Max_amount_of_Money["justify"] = "center"
        L_Max_amount_of_Money["text"] = "Max Amount of Money"
        L_Max_amount_of_Money.place(x=340,y=140,width=171,height=39)

        self.Input_Number_of_Volunteers=tk.Entry(self.Window)
        self.Input_Number_of_Volunteers["borderwidth"] = "1px"
        ft = tkFont.Font(family='Times',size=10)
        self.Input_Number_of_Volunteers["font"] = ft
        self.Input_Number_of_Volunteers["fg"] = "#333333"
        self.Input_Number_of_Volunteers["justify"] = "center"
        self.Input_Number_of_Volunteers["text"] = ""
        self.Input_Number_of_Volunteers.place(x=120,y=190,width=70,height=30)

        self.Input_Max_amount_of_Money=tk.Entry(self.Window)
        self.Input_Max_amount_of_Money["borderwidth"] = "1px"
        ft = tkFont.Font(family='Times',size=10)
        self.Input_Max_amount_of_Money["font"] = ft
        self.Input_Max_amount_of_Money["fg"] = "#333333"
        self.Input_Max_amount_of_Money["justify"] = "center"
        self.Input_Max_amount_of_Money["text"] = ""
        self.Input_Max_amount_of_Money.place(x=390,y=190,width=70,height=30)


        Choose_File_Button=tk.Button(self.Window)
        Choose_File_Button["bg"] = "#ff8c00"
        ft = tkFont.Font(family='Times',size=10)
        Choose_File_Button["font"] = ft
        Choose_File_Button["fg"] = "#000000"
        Choose_File_Button["justify"] = "center"
        Choose_File_Button["text"] = "Choose File"
        Choose_File_Button.place(x=260,y=260,width=70,height=25)
        Choose_File_Button["command"] = lambda: OpenFile(self)

        Start_Button=tk.Button(self.Window)
        Start_Button["bg"] = "#5fb878"
        ft = tkFont.Font(family='Times',size=10)
        Start_Button["font"] = ft
        Start_Button["fg"] = "#000000"
        Start_Button["justify"] = "center"
        Start_Button["text"] = "Start"
        Start_Button.place(x=260,y=340,width=70,height=25)
        Start_Button["command"] = lambda: Start(self)

        self.progressbar = ttk.Progressbar(self.Window) 
        self.progressbar.place(x=15, y=430, width=570)

        self.Window.bind("<Control-w>",lambda e:self.Window.destroy())


class Help_Window:
    def __init__(self):
        self.Window_Update_Database = Toplevel()
        self.Window_Update_Database.title("Help")
        self.Window_Update_Database.iconbitmap("./logo.ico")
        self.Window_Update_Database.geometry('800x600')

        self.Window_Update_Database.bind("<Control-w>",lambda e:self.Window_Update_Database.destroy())
        self.Window_Update_Database.grab_set()

class Window_Update_Database:
    def __init__(self):
        self.Window_Update_Database = Toplevel()
        self.Window_Update_Database.title("Update Database")
        self.Window_Update_Database.iconbitmap("./logo.ico")
        self.Window_Update_Database.geometry('300x150')

        self.Top_L_DB=tk.Label(self.Window_Update_Database)
        ft = tkFont.Font(family='Times',size=10)
        self.Top_L_DB["font"] = ft
        self.Top_L_DB["fg"] = "#333333"
        self.Top_L_DB["justify"] = "center"
        self.Top_L_DB["text"] = "Excel File :"
        self.Top_L_DB.place(x=10,y=20,width=120,height=35)

        self.Top_Button_Select_Database=tk.Button(self.Window_Update_Database)
        self.Top_Button_Select_Database["bg"] = "#e9e9ed"
        ft = tkFont.Font(family='Times',size=10)
        self.Top_Button_Select_Database["font"] = ft
        self.Top_Button_Select_Database["fg"] = "#000000"
        self.Top_Button_Select_Database["justify"] = "center"
        self.Top_Button_Select_Database["text"] = "Select File"
        self.Top_Button_Select_Database.place(x=160,y=20,width=96,height=30)
        self.Top_Button_Select_Database["command"] = self.Open_File_For_DB_Update

        self.Top_Label_Selected_File=tk.Label(self.Window_Update_Database)
        ft = tkFont.Font(family='Times',size=10)
        self.Top_Label_Selected_File["font"] = ft
        self.Top_Label_Selected_File["fg"] = "#333333"
        self.Top_Label_Selected_File["justify"] = "center"
        self.Top_Label_Selected_File["text"] = None
        self.Top_Label_Selected_File["wraplength"] = 150
        self.Top_Label_Selected_File.place(x=0,y=55,width=300,height=50)

        self.Top_Button_Start=tk.Button(self.Window_Update_Database)
        self.Top_Button_Start["bg"] = "#e9e9ed"
        ft = tkFont.Font(family='Times',size=10)
        self.Top_Button_Start["font"] = ft
        self.Top_Button_Start["fg"] = "#000000"
        self.Top_Button_Start["justify"] = "center"
        self.Top_Button_Start["text"] = "Update"
        self.Top_Button_Start.place(x=110,y=115,width=70,height=25)
        self.Top_Button_Start["command"] = lambda: self.Update_Database

        self.Window_Update_Database.bind("<Control-w>",lambda e:self.Window_Update_Database.destroy())
        self.Window_Update_Database.grab_set()


    def Open_File_For_DB_Update(self):
        filepath = askopenfilename(filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")])
        if not filepath:
            return None
        self.Top_Label_Selected_File.config(text=f"{filepath}")
        return filepath

    def Update_Database():

        return 0

if __name__ == "__main__":
    root = tk.Tk()
    window = MainWindow(root)
    root.mainloop()
    #root.protocol("WM_DELETE_WINDOW", on_closing)