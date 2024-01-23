
from threading import Thread
import requests,openpyxl,csv,json,sqlite3
from bs4 import BeautifulSoup
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import PatternFill
import tkinter as tk
import tkinter.font as tkFont
from tkinter.filedialog import askopenfilename
from tkinter import ttk,messagebox,Frame, Menu, Toplevel, Scrollbar, Canvas, Frame
from PIL import Image,ImageTk

DATABASE_LOCATION = 'Database.db'

LOGO_ICON_LOCATION = "_internal/assets/logo.ico"

#LOGO_ICON_LOCATION = "logo.ico"


def search_company_name(afm):
    url = "https://publicity.businessportal.gr/api/search"
    headers={
        "Host":"publicity.businessportal.gr",
        "User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:109.0) Gecko/20100101 Firefox/118.0",
        "Accept":"application/json, text/plain, */*",
        "Accept-Language":"en-US,en;q=0.5",
        "Accept-Encoding":"gzip, deflate, br",
        "Content-Type":"application/json",
        "Content-Length":"496",
        "Origin":"https://publicity.businessportal.gr",
        "Connection":"keep-alive",
        "Referer":"https://publicity.businessportal.gr/"
        }
    data={"dataToBeSent":{"inputField":afm,"city":None,"postcode":None,"legalType":[],"status":[],"suspension":[],"category":[],"specialCharacteristics":[],"employeeNumber":[],"armodiaGEMI":[],"kad":[],"recommendationDateFrom":None,"recommendationDateTo":None,"closingDateFrom":None,"closingDateTo":None,"alterationDateFrom":None,"alterationDateTo":None,"person":[],"personrecommendationDateFrom":None,"personrecommendationDateTo":None,"radioValue":"all","places":[]},"token":None,"language":"el"}
    response = requests.request('post',url, headers=headers,json=data)

    j = json.loads(response.text)
    lst = j['company']['hits']
    no_gemi = ""
    for rec in lst :
        suffix_gemi = rec["id"][-3:]
        if suffix_gemi in "000":
            return rec["name"]

def get_company_name(afm):
    base_url = "https://publicity.businessportal.gr/api/autocomplete/"
    url = base_url+afm

    headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36",
    "Origin": "https://publicity.businessportal.gr",
    "Referer": "https://publicity.businessportal.gr/",
    "Cookie": "next-i18next=el"
    }
    data = {"token":"null","language":"el"}
    response = requests.post(url, headers=headers)
    j = json.loads(response.text)
    lst = j['payload']['autocomplete']
    no_gemi = ""
    for rec in lst :
        if not rec.get("branchType"):
            no_gemi = str(rec['arGemi'])

    url = "https://publicity.businessportal.gr/api/company/details"
    payload = {"query":{"arGEMI":no_gemi},"language":"el"}
    response = requests.post(url, headers=headers,json=payload)
    j2 = json.loads(response.text)
    if j2['message'] in 'Company not found':
        return None
    company_name = j2['companyInfo']['payload']["company"]["name"]
    return company_name

def associate (Receipts,People,Max_Cost):
    N_Receipts = Receipts.copy()

    for p in People:

        for r in N_Receipts:
            if (r.Total > Max_Cost) or (r.Total< 0) or (p.Total + r.Total > Max_Cost) :
                continue
            p.Total += r.Total
            p.Array_Of_receipts.append(r)
            N_Receipts.remove(r)

class Reciept:
    def __init__(self,hmerominia,katharo_sunolo,sunolo_fpa,eidos_parastatikou,sunolo,arithmos_parastatikou,eidos_proiontos,afm,eponumia,url):
        self.Date_time = hmerominia
        self.Sum_of_net_price = katharo_sunolo
        self.Sum_of_fpa = sunolo_fpa
        self.R_type = eidos_parastatikou
        self.Total = float(sunolo.replace("€",""))
        self.Number = arithmos_parastatikou
        self.Type_of_goods = eidos_proiontos
        self.Company_afm = afm
        self.Company_Name = eponumia
        self.AADE_Url = url
    def set_id(self,_id):
        self.ID = _id
    
    def __repr__(self):
        return f"{self.ID}\t{self.Total}"

    def To_tuple(self):
        return (self.ID,self.Total)

    def fix_num(self):
        if type(self.Total) is type("1.1"):
            self.Total = float(self.Total.replace("€",""))

    def convert_to_record(self):
        return (2,self.Date_time,self.Sum_of_net_price,self.Sum_of_fpa," "," ",1," ",self.R_type,self.Number,self.Type_of_goods,"998727941","ΤΑΝΤΕΜ ΑΣΤΙΚΗ ΜΗ ΚΕΡΔΟΣΚΟΠΙΚΗ ΕΤΑΙΡΕΙΑ",1,self.Company_afm,self.Company_Name,"{:.2f}".format(self.Total),"=HYPERLINK(\""+self.AADE_Url+"\")")

class Person:
    def __init__(self,n,aor,tl):
        self.Name = n
        self.Array_Of_receipts = aor
        self.Total = tl
    
    def __repr__(self):
        return f"{self.Name} {self.Array_Of_receipts} {self.Total}"

def OpenFile(this):
    filepath = askopenfilename(filetypes=[("csv Files", "*.csv"), ("All Files", "*.*")])
    if not filepath:
        return None
    with open(filepath,'r',encoding='utf-8') as fl:
        doc = csv.reader(fl)
        for row in doc:
            Urls.append(row[4])

    this.L_File_Path.config(text=f"{filepath}")
    return Urls

def Execute(U,this):
    db = sqlite3.connect(DATABASE_LOCATION)
    cursor = db.cursor()
    Urls.pop(0)
    DATA_ARRAY = []
    errors = []
    Rs = []
    MAX_COST = float(this.Input_Max_amount_of_Money.get()) if this.Input_Max_amount_of_Money.get() != '' else 0
    No_of_candidates = int(this.Input_Number_of_Volunteers.get()) if this.Input_Number_of_Volunteers.get() != '' else 0

    for url in Urls:
        this.progressbar['value'] += 100/len(Urls)
        this.Window.update_idletasks()
        try:
            r = requests.get(url)
        except:
            errors.append(url)
            continue
        soup = BeautifulSoup(r.text,'lxml')
        table = soup.find('table',"info")
        if table is None:
            errors.append(url)
            continue
        tr = table.findAll('tr')
        DictArray = {}
        for row in tr:
            x = row.findAll('td')
            x = x[0].text , x[1].text
            DictArray[x[0]]= x[1]
        try:
            date_time = DictArray["Ημερομηνία, ώρα"].split(" ")[0]
            eidos_parastatikou = DictArray["Είδος παραστατικού"]
            receipt_number = DictArray["Αριθμός παραστατικού"]
            sum_net = str(round(float(DictArray['Καθαρή αξία Α'].replace("€ ","")) + float(DictArray['Καθαρή αξία Β'].replace("€ ","")) + float(DictArray['Καθαρή αξία Γ'].replace("€ ","")) + float(DictArray['Καθαρή αξία Δ'].replace("€ ","")) + float(DictArray['Καθαρή αξία Ε'].replace("€ ","")),2))+"€"
            sum_fpa = str(round(float(DictArray['ΦΠΑ Α'].replace("€ ","")) +float(DictArray['ΦΠΑ Β'].replace("€ ","")) + float(DictArray['ΦΠΑ Γ'].replace("€ ","")) + float(DictArray['ΦΠΑ Δ'].replace("€ ","")),2))+"€"
            sun_axia = str(round(float(DictArray[' Συνολική αξία '].rstrip(" ").replace("€ ","")),2))+"€"
            _afm = DictArray["ΑΦΜ εκδότη"]
        except:
            errors.append(url)
            continue
        cursor.execute(f"select Company_Name from Companies where AFM ='{_afm}';")
        company_name = cursor.fetchone()
        if company_name is not None:
            company_name = company_name[0]
            
        if company_name is None:
            company_name = search_company_name(_afm)
            if company_name is not None:
                cursor.execute(f"insert Into Companies(AFM,Company_name) values(?,?);",(_afm,company_name))

        if company_name is None:
            company_name = get_company_name(_afm)
            if company_name is not None:
                cursor.execute(f"insert Into Companies(AFM,Company_name) values(?,?);",(_afm,company_name))

        cursor.execute(f"select Product_Type from Companies where AFM ='{_afm}';")
        eidos_proiontos = cursor.fetchone()
        if eidos_proiontos is not None:
            eidos_proiontos = eidos_proiontos[0]
        url = url.rstrip('\x00')
        Reciept_Object = Reciept(date_time,sum_net,sum_fpa,eidos_parastatikou,sun_axia,receipt_number,eidos_proiontos,_afm,company_name,url)
        Rs.append(Reciept_Object)
        DATA_ARRAY.append(
            (2,date_time,sum_net,sum_fpa," "," ",1," ",eidos_parastatikou,receipt_number,eidos_proiontos,"998727941","ΤΑΝΤΕΜ ΑΣΤΙΚΗ ΜΗ ΚΕΡΔΟΣΚΟΠΙΚΗ ΕΤΑΙΡΕΙΑ",1,_afm,company_name,sun_axia,"=HYPERLINK(\""+url+"\")"))
    db.commit()


    Hashed_Receipts = {}

    for r in Rs:
        r.set_id(hash(r))
        r.fix_num()
        Hashed_Receipts[r.ID] = r

    items = [(r.ID,r.Total) for r in Rs]

    Rs.sort(key=lambda r:r.Total,reverse=True)

    Person_array = [Person(chr(i+65),[],0) for i in range(0,No_of_candidates)]
    Food_Receipts = [obj for obj in Rs if obj.Type_of_goods=="Τρόφιμα"]
    associate(Food_Receipts,Person_array,MAX_COST)
    Used_Reciepts = []
    for p in Person_array:
        for rec in p.Array_Of_receipts:
            Used_Reciepts.append(rec)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    for row in DATA_ARRAY:
        worksheet.append(row)

    worksheet.append(("","",""))
    for rows in worksheet.iter_rows(min_row=worksheet.max_row, max_row=worksheet.max_row, min_col=None):
        for cell in rows:
            cell.fill = PatternFill(start_color="9aec9f", end_color="9aec9f", fill_type = "solid")
    worksheet.append(("","","Grouped Reciepts","Food"))

    for p in Person_array:
        worksheet.append(("","Volunteer :"+p.Name,"Total","{:.2f}".format(p.Total)))
        for rec in p.Array_Of_receipts:
            worksheet.append(Hashed_Receipts[rec.ID].convert_to_record())
        worksheet.append(("","",""))

    for rows in worksheet.iter_rows(min_row=worksheet.max_row, max_row=worksheet.max_row, min_col=None):
        for cell in rows:
            cell.fill = PatternFill(start_color="619df9", end_color="619df9", fill_type = "solid")
    worksheet.append(("","","Grouped Reciepts","Non-Food"))
    del Person_array

    Person_array = [Person(chr(i+65),[],0) for i in range(0,No_of_candidates)]
    Remaining_Reciepts = [obj for obj in Rs if obj not in Used_Reciepts]
    associate(Remaining_Reciepts,Person_array,MAX_COST)

    for p in Person_array:
        worksheet.append(("","Volunteer :"+p.Name,"Total","{:.2f}".format(p.Total)))
        for rec in p.Array_Of_receipts:
            worksheet.append(Hashed_Receipts[rec.ID].convert_to_record())
        worksheet.append(("","",""))

    for rows in worksheet.iter_rows(min_row=worksheet.max_row, max_row=worksheet.max_row, min_col=None):
        for cell in rows:
            cell.fill = PatternFill(start_color="FF0000", end_color="ff0000", fill_type = "solid")

    worksheet.append((" "," "," "," "," "," "," "," "," "," "," "," "," "," "," "," "," ","ERRORS"))

    for e in errors:
        e = ILLEGAL_CHARACTERS_RE.sub(r'',e)
        worksheet.append((" "," "," "," "," "," "," "," "," "," "," "," "," "," "," "," "," ","=HYPERLINK(\""+e+"\")"))
    workbook.save("OUTPUT.xlsx")
    messagebox.showinfo("showinfo", "Task Completed") 

def Start(this):
    if Urls == []:
        messagebox.showinfo("showinfo", "Please Choose File")
        return None

    t = Thread(target=Execute,args=(Urls,this))
    t.start()
    if not t.is_alive() :
        messagebox.showinfo("showinfo", "Completed click ok to close")
        this.destroy()
    #Execute(Urls)

def Create_Label(Window,Font_Type='Times',Position = "center",Font_Size=10,Text_Color = "#333333", Label_text=""):
    Out_Label=tk.Label(Window)
    ft = tkFont.Font(family=Font_Type,size=Font_Size)
    Out_Label["font"] = ft
    Out_Label["fg"] = Text_Color
    Out_Label["justify"] = "center"
    Out_Label["text"] = Label_text
    Out_Label.place(x=60,y=40,width=130,height=32)
    return Out_Label  

def Create_Button(Window,Background_Color="#e9e9ed",Font_Type='Times',Font_Size=10,Text_Color = "#333333", Label_text=""):
    Out_Button=tk.Button(Window)
    Out_Button["bg"] = Background_Color
    ft = tkFont.Font(family='Times',size=Font_Size)
    Out_Button["font"] = ft
    Out_Button["fg"] = Text_Color
    Out_Button["justify"] = "center"
    Out_Button["text"] = Label_text
    return Out_Button

def Create_Input(Window,borderwidth="1px",Font_Type='Times',Font_Size=10,Text_Color = "#333333", Label_text=""):
    Out_InputField=tk.Entry(Window)
    Out_InputField["borderwidth"] = borderwidth
    ft = tkFont.Font(family=Font_Type,size=Font_Size)
    Out_InputField["font"] = ft
    Out_InputField["fg"] = Text_Color
    Out_InputField["justify"] = "center"
    Out_InputField["text"] = Label_text
    return Out_InputField


Urls = []
# region UI inits
class MainWindow:
    def __init__(self,root):
        self.Window = root
        self.Window.title("Receipt to XLS maker")
        width=600
        height=500
        screenwidth = self.Window.winfo_screenwidth()
        screenheight = self.Window.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height, (screenwidth - width) / 2, (screenheight - height) / 2)
        self.Window.geometry(alignstr)
        self.Window.resizable(width=False, height=False)

        self.Window.iconbitmap(LOGO_ICON_LOCATION)

        MenuBar = Menu(self.Window)
        self.Window.config(menu=MenuBar)

        Menu_File = Menu(MenuBar,tearoff=0)
        Menu_File.add_command(label="Search/Edit AFM", command=Window_Search_Or_Edit)
        Menu_File.add_command(label="Update Database", command=Window_Update_Database)
        Menu_File.add_separator()
        Menu_File.add_command(label="Exit", command=self.Window.destroy)

        MenuBar.add_cascade(label="File", menu=Menu_File)
        MenuBar.add_cascade(label="Help", command=Help_Window)

        self.L_Select_file = Create_Label(self.Window,Font_Size=13,Label_text="Selected File :")
        self.L_Select_file.place(x=60,y=40,width=130,height=32)

        self.L_File_Path= Create_Label(self.Window,Label_text="No file selected")
        self.L_File_Path.place(x=180,y=40,width=376,height=30)


        self.Choose_File_Button = Create_Button(self.Window,Background_Color="#ff8c00",Label_text="Choose File")
        self.Choose_File_Button.place(x=260,y=110,width=70,height=25)
        self.Choose_File_Button["command"] = lambda: OpenFile(self)

        self.L_Number_of_Volunteers = Create_Label(self.Window,Font_Size=12,Label_text="Number of Volunteers")
        self.L_Number_of_Volunteers.place(x=70,y=190,width=171,height=39)

        self.L_Max_amount_of_Money=Create_Label(self.Window,Font_Size=12,Label_text="Max Amount of Money")
        self.L_Max_amount_of_Money.place(x=340,y=190,width=171,height=39)

        self.Input_Number_of_Volunteers =  Create_Input(self.Window)
        self.Input_Number_of_Volunteers.place(x=120,y=240,width=70,height=30)

        self.Input_Max_amount_of_Money = Create_Input(self.Window)
        self.Input_Max_amount_of_Money.place(x=390,y=240,width=70,height=30)

        self.Start_Button = Create_Button(self.Window,Background_Color="#5fb878",Label_text="Start")
        self.Start_Button.place(x=260,y=330,width=70,height=25)
        self.Start_Button["command"] = lambda: Start(self)

        self.progressbar = ttk.Progressbar(self.Window) 
        self.progressbar.place(x=15, y=430, width=570)

        self.Window.bind("<Control-w>",lambda e:self.Window.destroy())

class Window_Update_Database:
    def __init__(self):
        self.Window_Update_Database = Toplevel()
        self.Window_Update_Database.title("Update Database")
        self.Window_Update_Database.iconbitmap(LOGO_ICON_LOCATION)
        width=300
        height=150
        screenwidth = self.Window_Update_Database.winfo_screenwidth()
        screenheight = self.Window_Update_Database.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height, (screenwidth - width) / 2, (screenheight - height) / 2)
        self.Window_Update_Database.geometry(alignstr)
        #self.Window_Update_Database.geometry('300x150')

        self.Top_L_DB =Create_Label(self.Window_Update_Database,Font_Size=10,Label_text="Excel File :")
        self.Top_L_DB.place(x=10,y=20,width=120,height=35)

        self.Top_Button_Select_Xlsx_File = Create_Button(self.Window_Update_Database,Label_text="Select File")
        self.Top_Button_Select_Xlsx_File.place(x=160,y=20,width=96,height=30)
        self.Top_Button_Select_Xlsx_File["command"] = self.Open_File_For_DB_Update

        self.Top_Label_Selected_File =Create_Label(self.Window_Update_Database,Font_Size=10)
        self.Top_L_DB.place(x=10,y=20,width=120,height=35)
        self.Top_Label_Selected_File["wraplength"] = 150
        self.Top_Label_Selected_File.place(x=0,y=55,width=300,height=50)


        self.Top_Button_Start = Create_Button(self.Window_Update_Database,Label_text="Update")
        self.Top_Button_Start.place(x=110,y=115,width=70,height=25)
        self.Top_Button_Start["command"] = self.Update_Database_Action

        self.Window_Update_Database.bind("<Control-w>",lambda e:self.Window_Update_Database.destroy())
        self.Window_Update_Database.grab_set()


    def Open_File_For_DB_Update(self):
        self.Excel_File_To_Insert = askopenfilename(filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")])
        if not self.Excel_File_To_Insert:
            self.Excel_File_To_Insert = None
            return None
        self.Top_Label_Selected_File.config(text=f"{self.Excel_File_To_Insert}")
        return self.Excel_File_To_Insert

    def Update_Database_Action(self):
        Data = []
        workbook = openpyxl.load_workbook(self.Excel_File_To_Insert)
        active_sheet = workbook.active
        for row in range(active_sheet.max_row):
            row_values = [x.value for x in active_sheet[row+1]]
            row_values = (str(row_values[2]),row_values[3],row_values[1])
            Data.append(row_values)
        workbook.close()
        Data.pop()  # Αφαίρεση της πρώτης γραμμης που ειναι ο Τίτλος
        with sqlite3.connect(DATABASE_LOCATION) as db:
            cursor = db.cursor()
            cursor.executemany("insert or replace into Companies (AFM, Company_Name,Product_Type) VALUES (?,?,?)", Data)
            db.commit()
        messagebox.showinfo("Task ended","Database Updated !")
        self.Window_Update_Database.destroy()
        return 0

class Window_Search_Or_Edit:
    def __init__(self):
        self.Window_Search_Or_Edit_Database = Toplevel()
        self.Window_Search_Or_Edit_Database.title("Search or edit record in database")
        self.Window_Search_Or_Edit_Database.iconbitmap(LOGO_ICON_LOCATION)
        width=310
        height=300
        screenwidth = self.Window_Search_Or_Edit_Database.winfo_screenwidth()
        screenheight = self.Window_Search_Or_Edit_Database.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height, (screenwidth - width) / 2, (screenheight - height) / 2)
        self.Window_Search_Or_Edit_Database.geometry(alignstr)
        #self.Window_Search_Or_Edit_Database.geometry('310x300')

        self.Window_Search_Or_Edit_Database.resizable(width=False, height=False)

        L_AFM = Create_Label(self.Window_Search_Or_Edit_Database,Label_text="ΑΦΜ :")
        L_AFM.place(x=0,y=20,width=70,height=25)

        self.Input_AFM = Create_Input(self.Window_Search_Or_Edit_Database,Font_Size=11,Label_text="Entry")
        #self.Input_AFM.bind("<Return>", Search_Inside_DB)
        self.Input_AFM.place(x=100,y=20,width=157,height=30)

        self.Button_Search_Database = Create_Button(self.Window_Search_Or_Edit_Database,Label_text="Αναζήτηση",Font_Size=10)
        self.Button_Search_Database.place(x=120,y=70,width=80,height=30)
        self.Button_Search_Database["command"] = self.Search_Inside_DB

        L_AFM_2 = Create_Label(self.Window_Search_Or_Edit_Database,Position="left",Font_Size=10,Label_text="ΑΦΜ :")
        L_AFM_2.place(x=10,y=130,width=70,height=25)

        self.L_Database_AFM = Create_Label(self.Window_Search_Or_Edit_Database,Font_Size=10,Label_text="")
        self.L_Database_AFM.place(x=80,y=130,width=116,height=30)

        L_Eponumia = Create_Label(self.Window_Search_Or_Edit_Database,Position="left",Font_Size=10,Label_text="Επωνυμία :")
        L_Eponumia.place(x=10,y=160,width=70,height=25)

        self.L_Database_Company = Create_Label(self.Window_Search_Or_Edit_Database,Font_Size=10,Label_text="")
        self.L_Database_Company["wraplength"] = 150
        self.L_Database_Company.place(x=85,y=160,width=150,height=30)

        L_Eidos= Create_Label(self.Window_Search_Or_Edit_Database,Position="left",Font_Size=10,Label_text="Είδος :")
        L_Eidos.place(x=10,y=190,width=70,height=25)

        self.L_Database_Type = Create_Label(self.Window_Search_Or_Edit_Database,Font_Size=10,Label_text="")
        self.L_Database_Type.place(x=80,y=190,width=116,height=30)

        L_Neo_Eidos = Create_Label(self.Window_Search_Or_Edit_Database,Font_Size=10,Label_text="Νέο Είδος :")
        L_Neo_Eidos.place(x=10,y=230,width=70,height=25)

        self.Input_New_Type = Create_Input(self.Window_Search_Or_Edit_Database,Font_Size=11)
        #self.Input_New_Type.bind("<Return>", self.Assign_New_Product_Type)
        self.Input_New_Type.place(x=100,y=230,width=100,height=20)

        self.Add_New_Button = Create_Button(self.Window_Search_Or_Edit_Database,Label_text="Εισαγωγή",Font_Size=10)
        self.Add_New_Button.place(x=210,y=230,width=80,height=30)
        self.Add_New_Button["command"] = self.Assign_New_Product_Type
        

        self.Window_Search_Or_Edit_Database.bind("<Control-w>",lambda e:self.Window_Search_Or_Edit_Database.destroy())
        self.Window_Search_Or_Edit_Database.grab_set()

    def Search_Inside_DB(self):

        with sqlite3.connect(DATABASE_LOCATION) as db:
            cursor = db.cursor()
            cursor.execute(f"select AFM,Company_Name,Product_Type from Companies where AFM = '{self.Input_AFM.get()}';")

            if res := cursor.fetchone():
                AFM,Company_Name,Product_Type = res
            else:
                AFM,Company_Name,Product_Type = "","",""
                messagebox.showerror("Not Found", "The ΑΦΜ doesn't exist !")

            self.L_Database_AFM.config(text=AFM)
            self.L_Database_Company.config(text=Company_Name)
            self.L_Database_Type.config(text=Product_Type)
       
        return 0

    def Assign_New_Product_Type(self):
        New_P_Type = self.Input_New_Type.get()
        
        if New_P_Type != "":
            with sqlite3.connect(DATABASE_LOCATION) as db:
                cursor = db.cursor()
                cursor.execute(f"Update Companies set Product_Type = '{New_P_Type}' where AFM = '{self.Input_AFM.get()}';")
                db.commit()
            self.Search_Inside_DB()
            self.Input_New_Type.delete(0,len(self.Input_New_Type.get()))
        return 0

class Help_Window:
    def __init__(self):
        self.Window_Help = Toplevel()
        self.Window_Help.title("Help")
        self.Window_Help.iconbitmap(LOGO_ICON_LOCATION)
        self.Window_Help.geometry('850x600')

        # Create a self.Canvas with a vertical scrollbar

        self.Canvas = Canvas(self.Window_Help, width=800, height=600)
        scrollbar = Scrollbar(self.Window_Help, command=self.Canvas.yview)
        frame = Frame(self.Canvas)

        Create_Label(self.Canvas, Font_Size=15, Label_text="Μορφές Αρχείων").pack()
        
        img= ImageTk.PhotoImage(file="Example_file_1.jpg")
        self.Canvas.create_image(10,10,anchor='center',image=img)

        self.Canvas.create_window(0,0,anchor='nw',window=frame)
        self.Canvas.update_idletasks()

        self.Canvas.configure(scrollregion=self.Canvas.bbox('all'), yscrollcommand=scrollbar.set)
                        
        self.Canvas.pack(fill='both', expand=True, side='left')
        scrollbar.pack(fill='y', side='right')

        self.Window_Help.bind("<Control-w>",lambda e:self.Window_Help.destroy())
        self.Canvas.bind_all("<MouseWheel>", lambda e:self.Canvas.yview_scroll(int(-1*(e.delta/120)), "units"))
        self.Window_Help.grab_set()

if __name__ == "__main__":
    root = tk.Tk()
    window = MainWindow(root)
    root.mainloop()